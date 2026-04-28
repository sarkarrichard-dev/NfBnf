from __future__ import annotations

import json
import sys
from collections.abc import Callable, Sequence
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

from nbnf.server import db
from nbnf.server.paths import REPO_ROOT, ml_data_roots

_SKIP_SUFFIXES = frozenset({".sqlite", ".db"})
_SKIP_NAMES = frozenset({"nbnf.sqlite"})
_SCAN_EXTENSIONS = frozenset({".csv", ".tsv", ".parquet", ".xlsx", ".xls"})
_DEFAULT_MAX_ROWS_PROFILE = 100_000
_MAX_SAMPLE_ROWS = 5

FileProgressFn = Callable[[str, str, int | None], None]
"""Callback(rel_path, phase, index): phase 'start'|'done'|'error'."""


def _rel_path(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return path.name


def _dataset_key_prefix(root: Path) -> str:
    """Stable path prefix for DB keys so files from different roots never collide."""
    try:
        rel = root.resolve().relative_to(REPO_ROOT.resolve())
        return str(rel).replace("\\", "/")
    except ValueError:
        r = root.resolve()
        return r.name or "dataset_root"


def _scan_one_root(
    root: Path,
    key_prefix: str,
    *,
    max_rows: int,
    on_file: FileProgressFn | None,
) -> tuple[int, list[str]]:
    processed = 0
    errors: list[str] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("."):
            continue
        # Excel temp/lock files
        if path.name.startswith("~$") and path.suffix.lower() in (".xlsx", ".xls"):
            continue
        if path.suffix.lower() in _SKIP_SUFFIXES or path.name in _SKIP_NAMES:
            continue
        if path.suffix.lower() not in _SCAN_EXTENSIONS:
            continue
        inner = _rel_path(path, root).replace("\\", "/")
        rel = f"{key_prefix}/{inner}" if key_prefix else inner
        rel_norm = rel.replace("\\", "/")
        if on_file:
            on_file(rel_norm, "start", processed)
        stat = path.stat()
        mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
        row: dict[str, Any] = {
            "rel_path": rel_norm,
            "format": path.suffix.lower().lstrip("."),
            "bytes": stat.st_size,
            "mtime": mtime,
            "rows_profiled": 0,
            "columns_json": "[]",
            "stats_json": "{}",
            "sample_head_json": "[]",
            "error": None,
        }
        try:
            df, ingest_meta = _read_frame(path, max_rows)
            row["rows_profiled"] = int(len(df))
            stats, cols, head = _profile_frame(df)
            if ingest_meta:
                stats = {**stats, **ingest_meta}
            row["stats_json"] = json.dumps(stats, default=str)
            row["columns_json"] = json.dumps(cols, default=str)
            row["sample_head_json"] = json.dumps(head, default=str)
            if on_file:
                on_file(rel_norm, "done", processed)
        except Exception as e:
            row["error"] = str(e)
            errors.append(f"{rel_norm}: {e}")
            if on_file:
                on_file(rel_norm, "error", processed)
        db.upsert_ml_dataset(row)
        processed += 1
    return processed, errors


def _read_excel_xlsx_head(path: Path, max_rows: int) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Stream first ``max_rows`` data rows from the first sheet — avoids loading multi‑GB workbooks into RAM."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "Excel (.xlsx) needs openpyxl: pip install openpyxl  (or pip install -e \".[data]\")"
        ) from e

    meta: dict[str, Any] = {
        "excel_profile": "openpyxl_read_only_first_sheet",
        "file_size_mb": round(path.stat().st_size / 1e6, 2),
    }
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        meta["sheet"] = ws.title
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return pd.DataFrame(), meta
        colnames = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
        n_cols = len(colnames)
        data: list[dict[str, Any]] = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            vals = list(row) if row is not None else []
            vals = (vals + [None] * n_cols)[:n_cols]
            data.append(dict(zip(colnames, vals)))
        df = pd.DataFrame(data)
        meta["rows_sampled_cap"] = max_rows
        return df, meta
    finally:
        wb.close()


def _read_parquet_head(path: Path, max_rows: int) -> tuple[pd.DataFrame, dict[str, Any]]:
    meta: dict[str, Any] = {"parquet_profile": "pandas_full_read_head", "file_size_mb": round(path.stat().st_size / 1e6, 2)}
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        df = pd.read_parquet(path)
        if len(df) > max_rows:
            df = df.iloc[:max_rows].copy()
        meta["parquet_profile"] = "pandas_full_read_head"
        return df, meta

    pf = pq.ParquetFile(path)
    batches: list[Any] = []
    total = 0
    batch_cap = min(65_536, max_rows)
    for batch in pf.iter_batches(batch_size=batch_cap):
        batches.append(batch)
        total += batch.num_rows
        if total >= max_rows:
            break
    if not batches:
        return pd.DataFrame(), {**meta, "parquet_profile": "pyarrow_empty"}
    table = pa.Table.from_batches(batches)
    if table.num_rows > max_rows:
        table = table.slice(0, max_rows)
    df = table.to_pandas()
    meta["parquet_profile"] = "pyarrow_iter_batches"
    meta["rows_sampled_cap"] = max_rows
    return df, meta


def _read_frame(path: Path, max_rows: int) -> tuple[pd.DataFrame, dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return pd.read_csv(path, nrows=max_rows, low_memory=False, encoding_errors="replace"), {}
    if suf == ".tsv":
        return (
            pd.read_csv(path, sep="\t", nrows=max_rows, low_memory=False, encoding_errors="replace"),
            {},
        )
    if suf == ".parquet":
        return _read_parquet_head(path, max_rows)
    if suf == ".xlsx":
        return _read_excel_xlsx_head(path, max_rows)
    if suf == ".xls":
        try:
            df = pd.read_excel(path, nrows=max_rows)
        except Exception as e:
            raise RuntimeError(
                "Legacy .xls: install xlrd or calamine, or convert to .xlsx — "
                f"original error: {e}"
            ) from e
        return df, {"excel_profile": "pandas_xls_first_sheet", "rows_sampled_cap": max_rows}
    raise ValueError(f"unsupported extension: {suf}")


def _profile_frame(df: pd.DataFrame) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    cols: list[dict[str, Any]] = []
    for c in df.columns:
        s = df[c]
        cols.append(
            {
                "name": str(c),
                "dtype": str(s.dtype),
                "null_pct": round(float(s.isna().mean()), 4),
                "nunique": int(s.nunique(dropna=True)),
            }
        )
    head_df = df.head(_MAX_SAMPLE_ROWS).copy()
    head = head_df.astype(object).where(pd.notna(head_df), None).to_dict(orient="records")
    stats = {
        "columns": len(df.columns),
        "memory_mb": round(float(df.memory_usage(deep=True).sum()) / 1e6, 3),
    }
    return stats, cols, head


def scan_and_ingest(
    roots: Sequence[Path] | None = None,
    *,
    max_rows_per_file: int = _DEFAULT_MAX_ROWS_PROFILE,
    on_file: FileProgressFn | None = None,
) -> dict[str, Any]:
    """
    Recursively walk each root (every subfolder), profile tabular files, upsert ``ml_datasets``.

    ``.xlsx`` uses openpyxl **read-only** streaming for the first sheet so multi‑GB files are not
    fully loaded into memory. ``.parquet`` uses PyArrow batch iteration when pyarrow is installed.
    """
    cap = max(1, int(max_rows_per_file))
    root_list = list(roots) if roots is not None else ml_data_roots()
    processed = 0
    errors: list[str] = []
    details: list[dict[str, Any]] = []
    for root in root_list:
        root = root.resolve()
        root.mkdir(parents=True, exist_ok=True)
        if not root.is_dir():
            msg = "not a directory"
            errors.append(f"{root}: {msg}")
            details.append({"root": str(root), "processed": 0, "errors": [msg]})
            continue
        prefix = _dataset_key_prefix(root)
        n, er = _scan_one_root(root, prefix, max_rows=cap, on_file=on_file)
        processed += n
        errors.extend(er)
        details.append({"root": str(root), "prefix": prefix, "processed": n, "errors": er})
    return {
        "roots": [str(r) for r in root_list],
        "processed": processed,
        "errors": errors,
        "details": details,
        "max_rows_per_file": cap,
    }


def default_cli_progress(rel_path: str, phase: str, _index: int | None) -> None:
    """Print each file to stderr (for long 5GB+ tree runs)."""
    if phase == "start":
        print(f"[ingest] {rel_path} …", file=sys.stderr, flush=True)
    elif phase == "done":
        print(f"[ingest] ok   {rel_path}", file=sys.stderr, flush=True)
    elif phase == "error":
        print(f"[ingest] FAIL {rel_path}", file=sys.stderr, flush=True)


def text_digest(*, max_files: int = 30, max_chars: int = 12_000) -> str:
    """Compact text for LLM context from last-ingested rows."""
    rows = db.fetch_ml_datasets(limit=max_files)
    chunks: list[str] = []
    for r in rows:
        if r.get("error"):
            chunks.append(f"- {r['rel_path']} ERROR: {r['error']}")
            continue
        try:
            cols = json.loads(r.get("columns_json") or "[]")
        except json.JSONDecodeError:
            cols = []
        col_preview = ", ".join(f"{c['name']}:{c['dtype']}" for c in cols[:12])
        if len(cols) > 12:
            col_preview += f" …(+{len(cols) - 12} cols)"
        chunks.append(
            f"- {r['rel_path']} rows≈{r['rows_profiled']} cols={len(cols)} "
            f"[{col_preview}] mtime={r.get('mtime', '')}"
        )
    text = "\n".join(chunks)
    return text[:max_chars]
